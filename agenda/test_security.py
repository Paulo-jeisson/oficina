from datetime import date, time
from io import BytesIO

from django.contrib.auth.models import User
from django.core.cache import cache
from django.core.files.uploadedfile import SimpleUploadedFile
from django.test import TestCase, override_settings
from django.urls import reverse
from openpyxl import load_workbook

from .forms import OficinaProfileForm
from .models import Booking, Oficina
from .security import safe_spreadsheet_value


class Sprint5SecurityTests(TestCase):
    def setUp(self):
        cache.clear()
        self.user = User.objects.create_user('tenant-a', password='correct', is_staff=True)
        self.other_user = User.objects.create_user('tenant-b', password='correct', is_staff=True)
        self.oficina = Oficina.objects.create(nome='Tenant A', dono=self.user)
        self.other = Oficina.objects.create(nome='Tenant B', dono=self.other_user)

    def create_booking(self, oficina, name='Cliente'):
        return Booking.objects.create(
            oficina=oficina,
            full_name=name,
            phone='11999999999',
            vehicle_brand='Fiat',
            vehicle_model='Uno',
            vehicle_year=2020,
            problem_description='Revisao',
            service_type='custom',
            duration_minutes=30,
            scheduled_date=date.today(),
            start_time=time(8),
            assigned_box=1,
        )

    def test_spreadsheet_values_neutralize_formula_prefixes(self):
        for value in ('=1+1', '+cmd', '-2+3', '@SUM(A1)', ' \t=1+1'):
            self.assertTrue(safe_spreadsheet_value(value).startswith("'"))
        self.assertEqual(safe_spreadsheet_value('Cliente normal'), 'Cliente normal')

    def test_booking_export_is_tenant_scoped_and_formula_safe(self):
        self.create_booking(self.oficina, '=HYPERLINK("https://invalid")')
        self.create_booking(self.other, 'SEGREDO OUTRA OFICINA')
        self.client.force_login(self.user)

        response = self.client.get(reverse('agenda:export_bookings_excel'))
        self.assertEqual(response.status_code, 200)
        workbook = load_workbook(BytesIO(response.content), data_only=False)
        values = [cell.value for row in workbook.active.iter_rows() for cell in row]

        self.assertIn("'=HYPERLINK(\"https://invalid\")", values)
        self.assertNotIn('SEGREDO OUTRA OFICINA', values)

    def test_foreign_booking_cannot_be_changed_or_deleted(self):
        foreign = self.create_booking(self.other)
        self.client.force_login(self.user)

        update = self.client.post(
            reverse('agenda:update_status'),
            {'booking_id': foreign.pk, 'status': Booking.Status.IN_PROGRESS},
        )
        delete = self.client.post(
            reverse('agenda:delete_booking'),
            {'booking_id': foreign.pk},
        )

        self.assertEqual(update.status_code, 404)
        self.assertEqual(delete.status_code, 404)
        self.assertTrue(Booking.objects.filter(pk=foreign.pk).exists())

    def test_logout_requires_post_and_external_referer_is_not_used(self):
        own = self.create_booking(self.oficina)
        self.client.force_login(self.user)
        self.assertEqual(self.client.get(reverse('agenda:logout')).status_code, 405)

        response = self.client.post(
            reverse('agenda:update_status'),
            {'booking_id': own.pk, 'status': Booking.Status.IN_PROGRESS},
            HTTP_REFERER='https://attacker.invalid/phishing',
        )
        self.assertRedirects(response, reverse('agenda:dashboard'))

    @override_settings(MAX_LOGO_UPLOAD_BYTES=16)
    def test_logo_rejects_oversized_and_fake_images(self):
        oversized = SimpleUploadedFile('logo.png', b'x' * 17, content_type='image/png')
        form = OficinaProfileForm(
            data={'nome': self.oficina.nome, 'mechanic_count': 1},
            files={'logo': oversized},
            instance=self.oficina,
        )
        self.assertFalse(form.is_valid())
        self.assertIn('logo', form.errors)

        fake = SimpleUploadedFile('logo.png', b'not-an-image', content_type='image/png')
        with override_settings(MAX_LOGO_UPLOAD_BYTES=1024):
            form = OficinaProfileForm(
                data={'nome': self.oficina.nome, 'mechanic_count': 1},
                files={'logo': fake},
                instance=self.oficina,
            )
            self.assertFalse(form.is_valid())
            self.assertIn('logo', form.errors)

    @override_settings(
        CACHES={'default': {'BACKEND': 'django.core.cache.backends.locmem.LocMemCache'}}
    )
    def test_login_is_limited_after_repeated_failures(self):
        login_url = reverse('agenda:login')
        for _ in range(10):
            response = self.client.post(
                login_url,
                {'username': self.user.username, 'password': 'wrong'},
                REMOTE_ADDR='203.0.113.10',
            )
            self.assertEqual(response.status_code, 200)

        response = self.client.post(
            login_url,
            {'username': self.user.username, 'password': 'wrong'},
            REMOTE_ADDR='203.0.113.10',
        )
        self.assertEqual(response.status_code, 429)
        self.assertIn('Retry-After', response)
