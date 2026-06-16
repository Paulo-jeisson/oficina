import json
import logging
import secrets
import re
from datetime import date, datetime, timedelta
from decimal import Decimal
from urllib.parse import quote
from django.conf import settings
from django.contrib import messages
from django.contrib.auth import login, logout
from django.contrib.auth.mixins import LoginRequiredMixin, UserPassesTestMixin
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from django.http import JsonResponse, HttpResponse
from django.shortcuts import get_object_or_404, redirect
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_http_methods
from django.utils.decorators import method_decorator
from django.views.generic import FormView, TemplateView, View
from django.urls import reverse, reverse_lazy
from django.utils import timezone
from django.db.models import Q, Sum, Count, Avg, F, Case, When, Value, DecimalField
from .asaas import AsaasClient, AsaasError
from .forms import (
    AssinaturaPaymentForm,
    BoxBlockForm,
    BookingForm,
    BookingDurationUpdateForm,
    OficinaProfileForm,
    OficinaSignupForm,
    OrdemServicoStatusForm,
    OrdemServicoServiceItemForm,
    OrdemServicoPartItemForm,
    OrdemServicoFinancialForm,
)
from .models import (
    Assinatura,
    BoxBlock,
    Booking,
    BUSINESS_MINUTES,
    BUSINESS_END,
    ServiceType,
    OrdemServico,
    OrdemServicoStatusHistory,
    OrdemServicoServiceItem,
    OrdemServicoPartItem,
    FinancialTransaction,
    CashFlowRecord,
    FinanceAudit,
    Oficina,
)
from .whatsapp import build_booking_owner_whatsapp_url
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

logger = logging.getLogger(__name__)


class PublicHomeView(TemplateView):
    template_name = 'agenda/site_home.html'


class PublicPlansView(TemplateView):
    template_name = 'agenda/site_plans.html'


class PublicContactView(TemplateView):
    template_name = 'agenda/site_contact.html'


class OficinaSignupView(FormView):
    template_name = 'agenda/signup.html'
    form_class = OficinaSignupForm
    success_url = reverse_lazy('agenda:dashboard')

    def form_valid(self, form):
        user = form.save()
        login(self.request, user)
        messages.success(self.request, 'Cadastro criado. Seu teste gratis de 20 dias ja comecou.')
        return super().form_valid(form)


class SubscriptionBlockedView(LoginRequiredMixin, TemplateView):
    template_name = 'agenda/subscription_blocked.html'
    login_url = reverse_lazy('agenda:login')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        assinatura = self.request.oficina.ensure_assinatura()
        assinatura.refresh_status()
        context['assinatura'] = assinatura
        context['payment_form'] = AssinaturaPaymentForm()
        return context


class SubscriptionDetailView(LoginRequiredMixin, TemplateView):
    template_name = 'agenda/subscription_detail.html'
    login_url = reverse_lazy('agenda:login')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        assinatura = self.request.oficina.ensure_assinatura()
        assinatura.refresh_status()
        today = timezone.localdate()
        trial_days_remaining = 0
        if assinatura.status == Assinatura.Status.TESTE and assinatura.trial_ends_at:
            trial_days_remaining = max((assinatura.trial_ends_at - today).days, 0)

        has_pending_charge = bool(
            assinatura.asaas_invoice_url
            and assinatura.asaas_payment_id
            and not assinatura.last_payment_at
        )

        context.update({
            'assinatura': assinatura,
            'payment_form': AssinaturaPaymentForm(),
            'trial_days_remaining': trial_days_remaining,
            'has_pending_charge': has_pending_charge,
            'today': today,
        })
        return context


class SubscriptionCancelView(LoginRequiredMixin, View):
    login_url = reverse_lazy('agenda:login')

    def post(self, request, *args, **kwargs):
        assinatura = request.oficina.ensure_assinatura()
        assinatura.refresh_status()
        if assinatura.status == Assinatura.Status.BLOQUEADO:
            messages.error(request, 'Assinatura bloqueada nao pode ser cancelada.')
            return redirect('agenda:subscription_detail')

        if assinatura.status != Assinatura.Status.CANCELADA:
            assinatura.status = Assinatura.Status.CANCELADA
            assinatura.save(update_fields=['status', 'updated_at'])

        due_date = assinatura.due_date.strftime('%d/%m/%Y') if assinatura.due_date else 'o vencimento atual'
        messages.success(request, f'Sua assinatura foi cancelada e permanecerá ativa até {due_date}.')
        return redirect('agenda:subscription_detail')


class AssinaturaPaymentCreateView(LoginRequiredMixin, View):
    login_url = reverse_lazy('agenda:login')

    def post(self, request, *args, **kwargs):
        fallback_url = request.POST.get('next') or request.META.get('HTTP_REFERER') or reverse('agenda:subscription_detail')
        form = AssinaturaPaymentForm(request.POST)
        if not form.is_valid():
            messages.error(request, 'Escolha uma forma de pagamento valida.')
            return redirect(fallback_url)

        assinatura = request.oficina.ensure_assinatura()
        has_pending_charge = bool(
            assinatura.asaas_invoice_url
            and assinatura.asaas_payment_id
            and not assinatura.last_payment_at
        )
        if has_pending_charge and request.POST.get('confirm_new_charge') != '1':
            messages.info(request, 'Ja existe uma cobranca gerada. Abra a ultima cobranca ou confirme a geracao de uma nova.')
            return redirect(fallback_url)

        try:
            payment = AsaasClient().create_payment(
                assinatura,
                form.cleaned_data['payment_method'],
                request=request,
            )
        except AsaasError as exc:
            messages.error(request, f'Nao foi possivel gerar a cobranca no Asaas: {exc}')
            return redirect(fallback_url)

        payment_url = payment.get('invoiceUrl') or assinatura.asaas_invoice_url
        if payment_url:
            return redirect(payment_url)

        messages.error(request, 'O Asaas criou a cobranca, mas nao retornou link de pagamento.')
        return redirect(fallback_url)


@method_decorator(csrf_exempt, name='dispatch')
class AsaasWebhookView(View):
    PAYMENT_PAID_EVENTS = {'PAYMENT_CONFIRMED', 'PAYMENT_RECEIVED'}

    def post(self, request, *args, **kwargs):
        if not self._is_valid_token(request):
            logger.warning('Webhook Asaas rejeitado por token invalido.')
            return JsonResponse({'detail': 'token invalido'}, status=403)

        try:
            payload = json.loads(request.body.decode('utf-8') or '{}')
        except json.JSONDecodeError:
            logger.warning('Webhook Asaas rejeitado por JSON invalido.')
            return JsonResponse({'detail': 'json invalido'}, status=400)

        event = payload.get('event')
        payment = payload.get('payment') or {}
        payment_id = payment.get('id', '')
        external_reference = payment.get('externalReference', '')
        logger.info(
            'Webhook Asaas recebido: event=%s payment_id=%s external_reference=%s status=%s billing_type=%s',
            event,
            payment_id,
            external_reference,
            payment.get('status', ''),
            payment.get('billingType', ''),
        )

        if event not in self.PAYMENT_PAID_EVENTS:
            logger.info('Webhook Asaas ignorado: event=%s payment_id=%s.', event, payment_id)
            return JsonResponse({'received': True, 'processed': False})

        assinatura = self._find_assinatura(payment_id, external_reference)
        if assinatura is None:
            logger.warning(
                'Webhook Asaas sem assinatura correspondente: event=%s payment_id=%s external_reference=%s.',
                event,
                payment_id,
                external_reference,
            )
            return JsonResponse({'received': True, 'processed': False})

        method = self._payment_method_from_asaas(payment)
        paid_at = self._payment_date_from_asaas(payment)
        previous_status = assinatura.status
        previous_due_date = assinatura.due_date
        already_processed = bool(
            payment_id
            and assinatura.asaas_payment_id == payment_id
            and assinatura.last_payment_at
        )

        if already_processed:
            logger.info(
                'Webhook Asaas ja processado anteriormente: assinatura_id=%s event=%s payment_id=%s.',
                assinatura.pk,
                event,
                payment_id,
            )
            self._register_audit_log(
                assinatura=assinatura,
                event=event,
                payment=payment,
                previous_status=previous_status,
                previous_due_date=previous_due_date,
                action='Webhook Asaas repetido para pagamento ja processado',
            )
            return JsonResponse({'received': True, 'processed': False, 'already_processed': True})

        if payment_id and not assinatura.asaas_payment_id:
            assinatura.asaas_payment_id = payment_id
            assinatura.save(update_fields=['asaas_payment_id', 'updated_at'])

        assinatura.mark_paid(payment_method=method, paid_at=paid_at)
        self._register_audit_log(
            assinatura=assinatura,
            event=event,
            payment=payment,
            previous_status=previous_status,
            previous_due_date=previous_due_date,
            action='Webhook Asaas confirmou pagamento de assinatura',
        )

        logger.info(
            'Webhook Asaas processado: assinatura_id=%s oficina_id=%s event=%s payment_id=%s status=%s due_date=%s.',
            assinatura.pk,
            assinatura.oficina_id,
            event,
            payment_id,
            assinatura.status,
            assinatura.due_date,
        )

        return JsonResponse({'received': True, 'processed': True})

    def _is_valid_token(self, request):
        expected_token = settings.ASAAS_WEBHOOK_TOKEN
        received_token = request.headers.get('asaas-access-token', '')
        return bool(expected_token and secrets.compare_digest(expected_token, received_token))

    def _find_assinatura(self, payment_id, external_reference):
        if payment_id:
            assinatura = Assinatura.objects.select_related('oficina').filter(asaas_payment_id=payment_id).first()
            if assinatura:
                return assinatura

        if external_reference.startswith('assinatura-'):
            assinatura_id = external_reference.replace('assinatura-', '', 1)
            return Assinatura.objects.select_related('oficina').filter(pk=assinatura_id).first()

        return None

    def _payment_method_from_asaas(self, payment):
        if payment.get('billingType') == 'CREDIT_CARD':
            return Assinatura.FormaPagamento.CARTAO_CREDITO
        return Assinatura.FormaPagamento.PIX

    def _payment_date_from_asaas(self, payment):
        date_value = payment.get('paymentDate') or payment.get('clientPaymentDate') or payment.get('confirmedDate')
        if not date_value:
            return None
        try:
            return datetime.strptime(date_value[:10], '%Y-%m-%d').date()
        except (TypeError, ValueError):
            logger.warning('Webhook Asaas recebeu data de pagamento invalida: %s.', date_value)
            return None

    def _register_audit_log(self, assinatura, event, payment, previous_status, previous_due_date, action):
        note = (
            f'Evento Asaas: {event}; '
            f'payment_id: {payment.get("id", "")}; '
            f'status Asaas: {payment.get("status", "")}; '
            f'forma: {payment.get("billingType", "")}; '
            f'valor: {payment.get("value", "")}; '
            f'status anterior: {previous_status}; '
            f'vencimento anterior: {previous_due_date}; '
            f'novo vencimento: {assinatura.due_date}.'
        )
        FinanceAudit.objects.create(
            oficina=assinatura.oficina,
            action=action,
            note=note,
        )


class AsaasConnectionTestView(LoginRequiredMixin, View):
    login_url = reverse_lazy('agenda:login')

    def get(self, request, *args, **kwargs):
        # Endpoint manual para validar Sandbox, chave e conectividade sem expor credenciais.
        try:
            data = AsaasClient().test_connection()
        except AsaasError as exc:
            return JsonResponse(
                {
                    'connected': False,
                    'environment': settings.ASAAS_BASE_URL,
                    'detail': str(exc),
                },
                status=exc.status_code or 502,
            )

        return JsonResponse({
            'connected': True,
            'environment': settings.ASAAS_BASE_URL,
            'customers_found': data.get('totalCount', 0),
        })


class AsaasPixTestChargeView(LoginRequiredMixin, UserPassesTestMixin, View):
    login_url = reverse_lazy('agenda:login')

    def test_func(self):
        return self.request.user.is_staff

    def get(self, request, *args, **kwargs):
        # Gera uma cobranca Pix pequena no Sandbox e retorna dados seguros para teste.
        oficina = getattr(request, 'oficina', None)
        if oficina is None:
            return JsonResponse({'detail': 'Oficina nao encontrada para o usuario logado.'}, status=400)

        assinatura = oficina.ensure_assinatura()
        try:
            result = AsaasClient().create_test_pix_charge(assinatura, value='5.00')
        except AsaasError as exc:
            return JsonResponse(
                {
                    'created': False,
                    'environment': settings.ASAAS_BASE_URL,
                    'detail': str(exc),
                },
                status=exc.status_code or 502,
            )

        payment = result['payment']
        qr_code = result['qr_code']
        encoded_image = qr_code.get('encodedImage', '')

        return JsonResponse({
            'created': True,
            'environment': settings.ASAAS_BASE_URL,
            'payment_id': payment.get('id'),
            'status': payment.get('status'),
            'value': payment.get('value'),
            'due_date': payment.get('dueDate') or result['due_date'],
            'invoice_url': payment.get('invoiceUrl', ''),
            'pix': {
                'qr_code_image_base64': encoded_image,
                'qr_code_image_data_url': f'data:image/png;base64,{encoded_image}' if encoded_image else '',
                'copy_paste_code': qr_code.get('payload', ''),
                'expiration_date': qr_code.get('expirationDate', ''),
            },
        })


@require_http_methods(['GET', 'POST'])
def logout_view(request):
    logout(request)
    return redirect('agenda:login')


class StaffRequiredMixin(LoginRequiredMixin, UserPassesTestMixin):
    login_url = reverse_lazy('agenda:login')

    def dispatch(self, request, *args, **kwargs):
        if request.user.is_authenticated and request.user.is_staff:
            self._ensure_user_oficina()
        return super().dispatch(request, *args, **kwargs)

    def _ensure_user_oficina(self):
        if hasattr(self.request.user, 'oficina'):
            return self.request.user.oficina
        oficina = Oficina.objects.create(
            dono=self.request.user,
            nome=f'Oficina de {self.request.user.get_username()}',
        )
        self.request.user.oficina = oficina
        return oficina

    def test_func(self):
        return self.request.user.is_staff and hasattr(self.request.user, 'oficina')

    def handle_no_permission(self):
        if self.request.user.is_authenticated:
            logout(self.request)
        return redirect(self.login_url)

    @property
    def oficina(self):
        return self._ensure_user_oficina()


class AppHomeView(StaffRequiredMixin, TemplateView):
    template_name = 'agenda/app_home.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['oficina'] = self.oficina
        return context


class MinhaOficinaView(StaffRequiredMixin, TemplateView):
    template_name = 'agenda/minha_oficina.html'

    def get_form(self):
        return OficinaProfileForm(
            self.request.POST or None,
            self.request.FILES or None,
            instance=self.oficina,
        )

    def post(self, request, *args, **kwargs):
        form = self.get_form()
        if form.is_valid():
            form.save()
            messages.success(request, 'Dados da oficina atualizados com sucesso.')
            return redirect('agenda:minha_oficina')

        messages.error(request, 'Nao foi possivel atualizar a oficina. Verifique os campos informados.')
        return self.render_to_response(self.get_context_data(form=form))

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        form = kwargs.get('form') or self.get_form()
        assinatura = self.oficina.ensure_assinatura()
        assinatura.refresh_status()
        public_booking_url = self.request.build_absolute_uri(
            reverse('agenda:public_booking', kwargs={'slug': self.oficina.slug})
        )
        context.update({
            'form': form,
            'oficina': self.oficina,
            'assinatura': assinatura,
            'public_booking_url': public_booking_url,
            'boxes_count': Booking.box_count_for_oficina(self.oficina),
        })
        return context


class BookingCreateView(FormView):
    template_name = 'agenda/booking_form.html'
    form_class = BookingForm
    success_url = reverse_lazy('agenda:booking_success')

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['oficina'] = getattr(self.request, 'oficina', None)
        return kwargs

    def form_valid(self, form):
        booking = form.save(commit=False)
        booking.oficina = getattr(self.request, 'oficina', None) or form.cleaned_data['oficina']
        assigned_box = Booking.find_first_available_box(
            booking.scheduled_date,
            booking.start_time,
            booking.duration_minutes,
            oficina=booking.oficina,
        )
        if assigned_box is None:
            form.add_error('start_time', 'O horario selecionado ja esta ocupado. Escolha outro horario.')
            return self.form_invalid(form)
        booking.assigned_box = assigned_box
        booking.save()

        return super().form_valid(form)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['service_durations'] = {
            'Troca de óleo': '30 minutos',
            'Analisar pessoalmente': '1 hora',
            'Revisão básica': '2 horas',
            'Revisão completa': '4 horas',
            'Diagnóstico eletrônico': '1 hora',
            'Serviço personalizado': 'Intervalo de tempo selecionável',
        }
        context['duration_help'] = {
            '30': '30 minutos',
            '60': '1 hora',
            '90': '1h 30min',
            '120': '2 horas',
            '240': '4 horas',
        }
        return context


class PublicBookingCreateView(BookingCreateView):
    template_name = 'agenda/public_booking_form.html'

    def dispatch(self, request, *args, **kwargs):
        self.public_oficina = get_object_or_404(Oficina, slug=kwargs.get('slug'))
        return super().dispatch(request, *args, **kwargs)

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['oficina'] = self.public_oficina
        return kwargs

    def get_success_url(self):
        return reverse('agenda:public_booking_success', kwargs={'slug': self.public_oficina.slug})

    def form_valid(self, form):
        booking = form.save(commit=False)
        booking.oficina = self.public_oficina
        assigned_box = Booking.find_first_available_box(
            booking.scheduled_date,
            booking.start_time,
            booking.duration_minutes,
            oficina=booking.oficina,
        )
        if assigned_box is None:
            form.add_error('start_time', 'O horario selecionado ja esta ocupado. Escolha outro horario.')
            return self.form_invalid(form)
        booking.assigned_box = assigned_box
        booking.save()
        self.request.session[f'public_booking_success_{self.public_oficina.slug}'] = booking.pk
        return super(BookingCreateView, self).form_valid(form)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['oficina'] = self.public_oficina
        context['public_booking_slug'] = self.public_oficina.slug
        return context


def get_public_booking_oficina(request, oficina_id=None, oficina_slug=None):
    if oficina_slug:
        try:
            return Oficina.objects.get(slug=oficina_slug)
        except Oficina.DoesNotExist:
            return None

    if oficina_id:
        try:
            return Oficina.objects.get(pk=oficina_id)
        except (Oficina.DoesNotExist, ValueError, TypeError):
            return None

    oficina = getattr(request, 'oficina', None)
    if oficina is not None:
        return oficina

    return Oficina.objects.order_by('nome').first()


class AvailableSlotsView(View):
    def get(self, request, *args, **kwargs):
        date_str = request.GET.get('date')
        duration_str = request.GET.get('duration')
        oficina_id = request.GET.get('oficina')
        oficina_slug = request.GET.get('oficina_slug')
        if not date_str or not duration_str:
            return JsonResponse({
                'slots': [],
                'message': 'Informe data e duração para exibir horários disponíveis.',
            })

        try:
            scheduled_date = datetime.strptime(date_str, '%Y-%m-%d').date()
            duration = int(duration_str)
        except (ValueError, TypeError):
            return JsonResponse({
                'slots': [],
                'message': 'Data ou duração inválida.',
            })

        if scheduled_date < date.today():
            return JsonResponse({
                'slots': [],
                'message': 'Não é possível agendar para datas passadas.',
            })

        oficina = get_public_booking_oficina(request, oficina_id, oficina_slug)
        if oficina is None:
            return JsonResponse({
                'slots': [],
                'message': 'Nenhuma oficina cadastrada para receber agendamentos.',
            })

        if Booking.available_minutes_for_date(scheduled_date, oficina=oficina) < duration:
            return JsonResponse({
                'slots': [],
                'message': 'Não há mais horários disponíveis para este dia.',
            })

        slots = Booking.available_start_times_for_date(scheduled_date, duration, oficina=oficina)
        if not slots:
            return JsonResponse({
                'slots': [],
                'message': 'Nenhum horário livre encontrado para a data e duração selecionadas.',
            })

        return JsonResponse({
            'slots': [slot.strftime('%H:%M') for slot in slots],
            'available_minutes': Booking.available_minutes_for_date(scheduled_date, oficina=oficina),
            'message': f'{len(slots)} horário(s) disponível(is).',
        })


class BookingSuccessView(TemplateView):
    template_name = 'agenda/booking_success.html'


class PublicBookingSuccessView(TemplateView):
    template_name = 'agenda/public_booking_success.html'

    def dispatch(self, request, *args, **kwargs):
        self.public_oficina = get_object_or_404(Oficina, slug=kwargs.get('slug'))
        return super().dispatch(request, *args, **kwargs)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        booking = self._get_last_public_booking()
        context['oficina'] = self.public_oficina
        context['booking'] = booking
        context['owner_whatsapp_url'] = build_booking_owner_whatsapp_url(booking) if booking else ''
        return context

    def _get_last_public_booking(self):
        booking_id = self.request.session.get(f'public_booking_success_{self.public_oficina.slug}')
        if not booking_id:
            return None
        return Booking.objects.filter(pk=booking_id, oficina=self.public_oficina).first()


class OrdemServicoListView(StaffRequiredMixin, TemplateView):
    template_name = 'agenda/os_list.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        query = self.request.GET.get('q', '').strip()
        status = self.request.GET.get('status', '')
        date_filter = self.request.GET.get('date', '')

        orders = OrdemServico.objects.select_related('booking').filter(oficina=self.oficina)
        if query:
            orders = orders.filter(
                Q(order_number__icontains=query)
                | Q(client_name__icontains=query)
                | Q(vehicle_brand__icontains=query)
                | Q(vehicle_model__icontains=query)
            )
        if status:
            orders = orders.filter(status=status)
        if date_filter:
            orders = orders.filter(scheduled_date=date_filter)

        orders = orders.order_by('-created_at')
        paginator = Paginator(orders, 5)
        page = self.request.GET.get('page')
        try:
            orders_page = paginator.page(page)
        except PageNotAnInteger:
            orders_page = paginator.page(1)
        except EmptyPage:
            orders_page = paginator.page(paginator.num_pages)

        filter_params = self.request.GET.copy()
        filter_params.pop('page', None)

        status_summary = {choice[0]: 0 for choice in OrdemServico.Status.choices}
        totals = OrdemServico.objects.filter(oficina=self.oficina).values('status').annotate(total=Count('id'))
        for item in totals:
            status_summary[item['status']] = item['total']

        context.update({
            'orders': orders_page,
            'orders_paginator': paginator,
            'total_orders': paginator.count,
            'query': query,
            'status_filter': status,
            'date_filter': date_filter,
            'filter_querystring': filter_params.urlencode(),
            'status_choices': OrdemServico.Status.choices,
            'status_summary': status_summary,
        })
        return context


class OrdemServicoDetailView(StaffRequiredMixin, TemplateView):
    template_name = 'agenda/os_detail_new.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        os_pk = self.kwargs.get('pk')
        ordem = get_object_or_404(OrdemServico, pk=os_pk, oficina=self.oficina)
        status_form = OrdemServicoStatusForm(initial={'status': ordem.status})
        financial_form = OrdemServicoFinancialForm(instance=ordem)
        service_form = OrdemServicoServiceItemForm()
        part_form = OrdemServicoPartItemForm()
        context.update({
            'ordem': ordem,
            'status_form': status_form,
            'financial_form': financial_form,
            'service_form': service_form,
            'part_form': part_form,
            'history': ordem.history.all(),
        })
        return context


class OrdemServicoServiceItemCreateView(StaffRequiredMixin, View):
    def post(self, request, pk, *args, **kwargs):
        ordem = get_object_or_404(OrdemServico, pk=pk, oficina=self.oficina)
        form = OrdemServicoServiceItemForm(request.POST)
        if form.is_valid():
            item = form.save(commit=False)
            item.oficina = self.oficina
            item.ordem_servico = ordem
            item.save()
            ordem.service_value = sum(i.total_price for i in ordem.service_items.all())
            ordem.save()
            return redirect('agenda:os_detail', pk=pk)
        return redirect('agenda:os_detail', pk=pk)


class OrdemServicoServiceItemDeleteView(StaffRequiredMixin, View):
    def post(self, request, pk, item_pk, *args, **kwargs):
        ordem = get_object_or_404(OrdemServico, pk=pk, oficina=self.oficina)
        item = get_object_or_404(OrdemServicoServiceItem, pk=item_pk, ordem_servico=ordem, oficina=self.oficina)
        item.delete()
        ordem.service_value = sum(i.total_price for i in ordem.service_items.all())
        ordem.save()
        return redirect('agenda:os_detail', pk=pk)


class OrdemServicoPartItemCreateView(StaffRequiredMixin, View):
    def post(self, request, pk, *args, **kwargs):
        ordem = get_object_or_404(OrdemServico, pk=pk, oficina=self.oficina)
        form = OrdemServicoPartItemForm(request.POST)
        if form.is_valid():
            item = form.save(commit=False)
            item.oficina = self.oficina
            item.ordem_servico = ordem
            item.save()
            ordem.parts_value = sum(i.total_price for i in ordem.part_items.all())
            ordem.save()
            return redirect('agenda:os_detail', pk=pk)
        return redirect('agenda:os_detail', pk=pk)


class OrdemServicoPartItemDeleteView(StaffRequiredMixin, View):
    def post(self, request, pk, item_pk, *args, **kwargs):
        ordem = get_object_or_404(OrdemServico, pk=pk, oficina=self.oficina)
        item = get_object_or_404(OrdemServicoPartItem, pk=item_pk, ordem_servico=ordem, oficina=self.oficina)
        item.delete()
        ordem.parts_value = sum(i.total_price for i in ordem.part_items.all())
        ordem.save()
        return redirect('agenda:os_detail', pk=pk)


class OrdemServicoFinancialUpdateView(StaffRequiredMixin, View):
    def post(self, request, pk, *args, **kwargs):
        ordem = get_object_or_404(OrdemServico, pk=pk, oficina=self.oficina)
        form = OrdemServicoFinancialForm(request.POST, instance=ordem)
        if form.is_valid():
            form.save()
        return redirect('agenda:os_detail', pk=pk)


class OrdemServicoCreateFromBookingView(StaffRequiredMixin, View):
    def get(self, request, booking_id, *args, **kwargs):
        booking = get_object_or_404(Booking, pk=booking_id, oficina=self.oficina)
        if hasattr(booking, 'ordem_servico'):
            return redirect('agenda:os_detail', pk=booking.ordem_servico.pk)

        ordem = OrdemServico.objects.create(
            oficina=self.oficina,
            booking=booking,
            client_name=booking.full_name,
            phone=booking.phone,
            vehicle_brand=booking.vehicle_brand,
            vehicle_model=booking.vehicle_model,
            vehicle_year=booking.vehicle_year,
            plate='',
            problem_description=booking.problem_description,
            duration_minutes=booking.duration_minutes,
            scheduled_date=booking.scheduled_date,
            observations='',
            status=OrdemServico.Status.RECEBIDO,
        )
        return redirect('agenda:os_detail', pk=ordem.pk)


class OrdemServicoStatusUpdateView(StaffRequiredMixin, View):
    def post(self, request, pk, *args, **kwargs):
        ordem = get_object_or_404(OrdemServico, pk=pk, oficina=self.oficina)
        form = OrdemServicoStatusForm(request.POST)
        if form.is_valid():
            new_status = form.cleaned_data['status']
            note = form.cleaned_data['note']
            if new_status != ordem.status:
                ordem.status = new_status
                ordem.save(user=request.user, note=note)
        return redirect('agenda:os_detail', pk=ordem.pk)


class OrdemServicoPrintView(StaffRequiredMixin, TemplateView):
    template_name = 'agenda/os_print.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        os_pk = self.kwargs.get('pk')
        ordem = get_object_or_404(OrdemServico, pk=os_pk, oficina=self.oficina)
        context['ordem'] = ordem
        return context


class DashboardView(StaffRequiredMixin, TemplateView):
    template_name = 'agenda/dashboard.html'

    @staticmethod
    def _normalize_whatsapp_phone(phone):
        digits = re.sub(r'\D+', '', phone or '')
        if not digits:
            return ''
        if digits.startswith('55'):
            return digits if len(digits) in (12, 13) else ''
        if len(digits) in (10, 11):
            return f'55{digits}'
        return ''

    def _build_pending_whatsapp_items(self):
        pending_bookings = (
            Booking.objects.filter(oficina=self.oficina, status=Booking.Status.SCHEDULED)
            .order_by('-created_at')[:5]
        )
        items = []
        for booking in pending_bookings:
            phone = self._normalize_whatsapp_phone(booking.phone)
            message = (
                f'Olá, {booking.full_name}! Recebemos seu agendamento na {self.oficina.nome} '
                f'para o dia {booking.scheduled_date.strftime("%d/%m/%Y")} às '
                f'{booking.start_time.strftime("%H:%M")}. Serviço: {booking.get_service_type_display()}. '
                'Podemos confirmar?'
            )
            items.append({
                'booking': booking,
                'whatsapp_url': f'https://wa.me/{phone}?text={quote(message)}' if phone else '',
            })
        return items

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        query = self.request.GET.get('q', '').strip()
        status = self.request.GET.get('status', '')
        date_filter = self.request.GET.get('date', '')

        bookings = Booking.objects.filter(oficina=self.oficina)
        if query:
            bookings = bookings.filter(
                Q(full_name__icontains=query)
                | Q(vehicle_brand__icontains=query)
                | Q(vehicle_model__icontains=query)
            )
        if status:
            bookings = bookings.filter(status=status)
        if date_filter:
            bookings = bookings.filter(scheduled_date=date_filter)

        bookings = bookings.order_by('-scheduled_date', 'start_time', 'full_name')
        
        # Implementar paginação
        paginator = Paginator(bookings, 5)
        page = self.request.GET.get('page')
        try:
            bookings_page = paginator.page(page)
        except PageNotAnInteger:
            bookings_page = paginator.page(1)
        except EmptyPage:
            bookings_page = paginator.page(paginator.num_pages)

        calendar_start = date.today()
        calendar_days = [calendar_start + timedelta(days=offset) for offset in range(7)]
        daily_summary = []
        for current_date in calendar_days:
            booked = Booking.booked_minutes_for_date(current_date, oficina=self.oficina)
            daily_capacity = BUSINESS_MINUTES * Booking.box_count_for_oficina(self.oficina)
            occupancy = round((booked / daily_capacity) * 100) if booked and daily_capacity else 0
            daily_summary.append({
                'date': current_date,
                'booked': booked,
                'available': Booking.available_minutes_for_date(current_date, oficina=self.oficina),
                'occupancy': occupancy,
                'intervals': Booking.occupied_intervals_for_date(current_date, oficina=self.oficina),
            })

        status_counts = Booking.objects.filter(oficina=self.oficina).values('status').annotate(total=Sum('duration_minutes'))
        status_map = {item['status']: item['total'] for item in status_counts}
        in_progress_count = Booking.objects.filter(oficina=self.oficina, status=Booking.Status.IN_PROGRESS).count()
        filter_params = self.request.GET.copy()
        filter_params.pop('page', None)

        selected_schedule = []
        if date_filter:
            try:
                selected_date = datetime.strptime(date_filter, '%Y-%m-%d').date()
                selected_schedule = Booking.occupied_intervals_for_date(selected_date, oficina=self.oficina)
            except ValueError:
                selected_schedule = []

        context.update({
            'bookings': bookings_page,
            'paginator': paginator,
            'total_bookings': paginator.count,
            'public_booking_url': self.request.build_absolute_uri(
                reverse('agenda:public_booking', kwargs={'slug': self.oficina.slug})
            ),
            'calendar_days': calendar_days,
            'daily_summary': daily_summary,
            'status_totals': status_map,
            'in_progress_count': in_progress_count,
            'query': query,
            'status_filter': status,
            'date_filter': date_filter,
            'filter_querystring': filter_params.urlencode(),
            'selected_schedule': selected_schedule,
            'status_choices': Booking.Status.choices,
            'pending_whatsapp_bookings': self._build_pending_whatsapp_items(),
        })
        return context


class BoxesPanelView(StaffRequiredMixin, TemplateView):
    template_name = 'agenda/boxes_panel.html'

    def get_selected_date(self):
        date_value = self.request.GET.get('date') or self.request.POST.get('selected_date')
        if date_value:
            try:
                return datetime.strptime(date_value, '%Y-%m-%d').date()
            except (TypeError, ValueError):
                messages.error(self.request, 'Data invalida. Exibindo atendimentos de hoje.')
        return timezone.localdate()

    def post(self, request, *args, **kwargs):
        selected_date = request.POST.get('selected_date') or timezone.localdate().isoformat()
        redirect_url = f'{reverse("agenda:boxes_panel")}?date={selected_date}'
        action = request.POST.get('action')

        if action == 'create_block':
            return self.create_block(request, redirect_url)

        if action == 'delete_block':
            return self.delete_block(request, redirect_url)

        return self.update_booking_duration(request, redirect_url)

    def create_block(self, request, redirect_url):
        form = BoxBlockForm(request.POST, oficina=self.oficina)
        if not form.is_valid():
            message = self._first_form_error(form) or 'Nao foi possivel bloquear este Box.'
            messages.error(request, message)
            return redirect(redirect_url)

        block = form.save(commit=False)
        block.oficina = self.oficina
        block.save()
        messages.success(request, f'{block.box_label} bloqueado com sucesso.')
        return redirect(redirect_url)

    def delete_block(self, request, redirect_url):
        block = get_object_or_404(BoxBlock, pk=request.POST.get('block_id'), oficina=self.oficina)
        box_label = block.box_label
        block.delete()
        messages.success(request, f'{box_label} desbloqueado com sucesso.')
        return redirect(redirect_url)

    def update_booking_duration(self, request, redirect_url):
        form = BookingDurationUpdateForm(request.POST)
        if not form.is_valid():
            messages.error(request, 'Informe uma duracao valida para o atendimento.')
            return redirect(redirect_url)

        booking = get_object_or_404(
            Booking,
            pk=form.cleaned_data['booking_id'],
            oficina=self.oficina,
        )
        new_duration = form.cleaned_data['duration_minutes']
        new_end_minutes = Booking._time_to_minutes(booking.start_time) + new_duration

        if new_end_minutes > Booking._time_to_minutes(BUSINESS_END):
            messages.error(request, 'A duracao informada ultrapassa o horario de funcionamento.')
            return redirect(redirect_url)

        has_conflict = Booking.box_has_conflict(
            booking.scheduled_date,
            booking.start_time,
            new_duration,
            oficina=self.oficina,
            assigned_box=booking.assigned_box,
            exclude_pk=booking.pk,
        )
        if has_conflict:
            messages.error(request, 'Nao foi possivel alterar: a nova duracao invade outro atendimento neste Box.')
            return redirect(redirect_url)

        booking.duration_minutes = new_duration
        booking.save(update_fields=['duration_minutes', 'updated_at'])
        messages.success(request, f'Duracao atualizada para {booking.assigned_box_label}.')
        return redirect(redirect_url)

    def _first_form_error(self, form):
        for field_errors in form.errors.values():
            if field_errors:
                return field_errors[0]
        return ''

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        selected_date = self.get_selected_date()
        bookings = Booking.objects.filter(
            oficina=self.oficina,
            scheduled_date=selected_date,
        ).exclude(status=Booking.Status.CANCELED).order_by('assigned_box', 'start_time', 'full_name')

        bookings_by_box = {box_index: [] for box_index in Booking.box_indexes_for_oficina(self.oficina)}
        for booking in bookings:
            if booking.assigned_box not in bookings_by_box:
                continue
            bookings_by_box[booking.assigned_box].append(booking)

        active_blocks = BoxBlock.objects.filter(
            oficina=self.oficina,
            end_datetime__gte=timezone.now(),
        ).order_by('start_datetime')
        blocks_by_box = {box_index: [] for box_index in bookings_by_box}
        for block in active_blocks:
            if block.box_number in blocks_by_box:
                blocks_by_box[block.box_number].append(block)

        boxes = [
            {
                'index': box_index,
                'label': self.oficina.get_box_label(box_index),
                'bookings': bookings_by_box[box_index],
                'blocks': blocks_by_box[box_index],
            }
            for box_index in bookings_by_box
        ]

        context.update({
            'selected_date': selected_date,
            'selected_date_value': selected_date.isoformat(),
            'boxes': boxes,
            'block_form': BoxBlockForm(oficina=self.oficina),
            'duration_form': BookingDurationUpdateForm(),
            'duration_choices': BookingDurationUpdateForm.base_fields['duration_minutes'].choices,
        })
        return context


class ExportBookingsExcelView(StaffRequiredMixin, View):
    def get(self, request, *args, **kwargs):
        query = request.GET.get('q', '').strip()
        status = request.GET.get('status', '')
        date_filter = request.GET.get('date', '')

        bookings = Booking.objects.filter(oficina=self.oficina)
        if query:
            bookings = bookings.filter(
                Q(full_name__icontains=query)
                | Q(vehicle_brand__icontains=query)
                | Q(vehicle_model__icontains=query)
            )
        if status:
            bookings = bookings.filter(status=status)
        if date_filter:
            bookings = bookings.filter(scheduled_date=date_filter)

        bookings = bookings.order_by('-scheduled_date', 'start_time', 'full_name')

        # Criar workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Agendamentos"

        # Estilos
        header_font = Font(bold=True, color="FFFFFF", size=12)
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Cabeçalhos
        headers = [
            'Nome',
            'Telefone',
            'Marca',
            'Modelo',
            'Ano',
            'Problema',
            'Tempo Estimado',
            'Data',
            'Horário',
            'Status'
        ]
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border

        # Adicionar dados
        for row_num, booking in enumerate(bookings, 2):
            data = [
                booking.full_name,
                booking.phone,
                booking.vehicle_brand,
                booking.vehicle_model,
                booking.vehicle_year,
                booking.problem_description,
                booking.duration_label,
                booking.scheduled_date.strftime('%d/%m/%Y'),
                booking.start_time.strftime('%H:%M'),
                booking.status_label,
            ]
            
            for col_num, value in enumerate(data, 1):
                cell = ws.cell(row=row_num, column=col_num)
                cell.value = value
                cell.border = border
                cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

        # Ajustar largura das colunas
        column_widths = [20, 15, 15, 15, 8, 30, 18, 12, 10, 15]
        for col_num, width in enumerate(column_widths, 1):
            ws.column_dimensions[chr(64 + col_num)].width = width

        # Preparar resposta
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="agendamentos.xlsx"'
        wb.save(response)
        
        return response


class FinanceDashboardView(StaffRequiredMixin, TemplateView):
    template_name = 'agenda/finance_dashboard.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        start_date = self.request.GET.get('start_date', '')
        end_date = self.request.GET.get('end_date', '')
        client = self.request.GET.get('client', '').strip()
        vehicle = self.request.GET.get('vehicle', '').strip()
        mechanic = self.request.GET.get('mechanic', '').strip()
        status = self.request.GET.get('status', '')
        filter_params = self.request.GET.copy()
        filter_params.pop('page', None)
        filter_querystring = filter_params.urlencode()

        orders = OrdemServico.objects.filter(oficina=self.oficina, status=OrdemServico.Status.COMPLETED)
        if start_date:
            orders = orders.filter(completed_at__date__gte=start_date)
        if end_date:
            orders = orders.filter(completed_at__date__lte=end_date)
        if client:
            orders = orders.filter(client_name__icontains=client)
        if vehicle:
            orders = orders.filter(Q(vehicle_brand__icontains=vehicle) | Q(vehicle_model__icontains=vehicle) | Q(plate__icontains=vehicle))
        if mechanic:
            orders = orders.filter(mechanic_name__icontains=mechanic)
        if status:
            orders = orders.filter(status=status)

        revenue_day = orders.filter(completed_at__date=date.today()).aggregate(total=Sum('total_value'))['total'] or 0
        revenue_week = orders.filter(completed_at__date__gte=date.today() - timedelta(days=7)).aggregate(total=Sum('total_value'))['total'] or 0
        revenue_month = orders.filter(completed_at__month=date.today().month, completed_at__year=date.today().year).aggregate(total=Sum('total_value'))['total'] or 0
        revenue_year = orders.filter(completed_at__year=date.today().year).aggregate(total=Sum('total_value'))['total'] or 0
        total_completed_os = orders.count()
        unique_clients = orders.values('client_name').distinct().count()
        unique_vehicles = orders.values('vehicle_brand', 'vehicle_model', 'plate').distinct().count()
        ticket_mean = orders.aggregate(avg=Avg('total_value'))['avg'] or 0

        revenue_total = orders.aggregate(total=Sum('total_value'))['total'] or 0
        services_done = orders.count()
        parts_total = orders.aggregate(total=Sum('parts_value'))['total'] or 0
        # removed estimated_profit aggregation

        top_services = OrdemServicoServiceItem.objects.filter(oficina=self.oficina).values('description').annotate(total=Sum(F('quantity'))).order_by('-total')[:5]
        top_parts = OrdemServicoPartItem.objects.filter(oficina=self.oficina).values('description').annotate(total=Sum(F('quantity'))).order_by('-total')[:5]

        daily_series = []
        for offset in range(7):
            day = date.today() - timedelta(days=6 - offset)
            amount = orders.filter(completed_at__date=day).aggregate(total=Sum('total_value'))['total'] or 0
            daily_series.append({'date': day.strftime('%d/%m'), 'amount': float(amount)})

        monthly_series = []
        for month_offset in range(6, -1, -1):
            current = date.today().replace(day=1) - timedelta(days=month_offset * 30)
            month_label = current.strftime('%b/%Y')
            amount = orders.filter(completed_at__month=current.month, completed_at__year=current.year).aggregate(total=Sum('total_value'))['total'] or 0
            monthly_series.append({'month': month_label, 'amount': float(amount)})

        service_vs_parts = [
            {'label': 'Serviços', 'value': float(orders.aggregate(total=Sum('service_value'))['total'] or 0)},
            {'label': 'Peças', 'value': float(parts_total)},
        ]

        transaction_filters = FinancialTransaction.objects.select_related('ordem_servico').filter(
            oficina=self.oficina,
            ordem_servico__isnull=False
        )
        if start_date:
            transaction_filters = transaction_filters.filter(due_date__gte=start_date)
        if end_date:
            transaction_filters = transaction_filters.filter(due_date__lte=end_date)
        if client:
            transaction_filters = transaction_filters.filter(ordem_servico__client_name__icontains=client)
        if vehicle:
            transaction_filters = transaction_filters.filter(
                Q(ordem_servico__vehicle_brand__icontains=vehicle)
                | Q(ordem_servico__vehicle_model__icontains=vehicle)
                | Q(ordem_servico__plate__icontains=vehicle)
            )
        if mechanic:
            transaction_filters = transaction_filters.filter(ordem_servico__mechanic_name__icontains=mechanic)
        if status:
            transaction_filters = transaction_filters.filter(ordem_servico__status=status)

        pending_alerts = transaction_filters.filter(status='pending', due_date__lt=date.today()).count()
        receivables = transaction_filters.filter(transaction_type='receivable').order_by('-due_date')
        receivables_paginator = Paginator(receivables, 5)
        page = self.request.GET.get('page')
        try:
            receivables_page = receivables_paginator.page(page)
        except PageNotAnInteger:
            receivables_page = receivables_paginator.page(1)
        except EmptyPage:
            receivables_page = receivables_paginator.page(receivables_paginator.num_pages)
        payables = transaction_filters.filter(transaction_type='payable').order_by('-due_date')[:15]
        cashflow_balance = CashFlowRecord.objects.filter(oficina=self.oficina).aggregate(balance=Sum(Case(When(entry_type='inflow', then=F('amount')), When(entry_type='outflow', then=F('amount') * Value(-1)), default=Value(0), output_field=DecimalField())))['balance'] or 0

        context.update({
            'start_date': start_date,
            'end_date': end_date,
            'client': client,
            'vehicle': vehicle,
            'mechanic': mechanic,
            'status_filter': status,
            'filter_querystring': filter_querystring,
            'status_choices': OrdemServico.Status.choices,
            'revenue_day': revenue_day,
            'revenue_week': revenue_week,
            'revenue_month': revenue_month,
            'revenue_year': revenue_year,
            'total_completed_os': total_completed_os,
            'ticket_mean': ticket_mean,
            'unique_clients': unique_clients,
            'unique_vehicles': unique_vehicles,
            'revenue_total': revenue_total,
            'services_done': services_done,
            'parts_total': parts_total,
            'top_services': list(top_services),
            'top_parts': list(top_parts),
            'daily_series': daily_series,
            'monthly_series': monthly_series,
            'service_vs_parts': service_vs_parts,
            'pending_alerts': pending_alerts,
            'receivables': receivables_page,
            'receivables_paginator': receivables_paginator,
            'total_receivables': receivables_paginator.count,
            'payables': payables,
            'cashflow_balance': cashflow_balance,
        })
        return context


class FinanceExportExcelView(StaffRequiredMixin, View):
    def get(self, request, *args, **kwargs):
        start_date = request.GET.get('start_date', '')
        end_date = request.GET.get('end_date', '')
        client = request.GET.get('client', '').strip()
        vehicle = request.GET.get('vehicle', '').strip()
        mechanic = request.GET.get('mechanic', '').strip()
        status = request.GET.get('status', '')

        orders = OrdemServico.objects.filter(oficina=self.oficina, status=OrdemServico.Status.COMPLETED)
        if start_date:
            orders = orders.filter(completed_at__date__gte=start_date)
        if end_date:
            orders = orders.filter(completed_at__date__lte=end_date)
        if client:
            orders = orders.filter(client_name__icontains=client)
        if vehicle:
            orders = orders.filter(Q(vehicle_brand__icontains=vehicle) | Q(vehicle_model__icontains=vehicle) | Q(plate__icontains=vehicle))
        if mechanic:
            orders = orders.filter(mechanic_name__icontains=mechanic)
        if status:
            orders = orders.filter(status=status)

        wb = Workbook()
        ws = wb.active
        ws.title = 'Financeiro'

        headers = [
            'OS', 'Cliente', 'Veículo', 'Placa', 'Mecânico', 'Data Conclusão', 'Serviços', 'Peças', 'Total', 'Status', 'Pagamento'
        ]
        header_font = Font(bold=True, color='FFFFFF', size=12)
        header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = header_alignment

        for row_num, ordem in enumerate(orders, 2):
            row = [
                ordem.order_number,
                ordem.client_name,
                f'{ordem.vehicle_brand} {ordem.vehicle_model}',
                ordem.plate,
                ordem.mechanic_name,
                ordem.completed_at.strftime('%d/%m/%Y %H:%M') if ordem.completed_at else '',
                float(ordem.service_value),
                float(ordem.parts_value),
                float(ordem.total_value),
                ordem.status_label,
                ordem.payment_method,
            ]
            for col_num, value in enumerate(row, 1):
                cell = ws.cell(row=row_num, column=col_num)
                cell.value = value
                cell.border = border
                cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)

        widths = [14, 25, 20, 12, 18, 18, 12, 12, 12, 14, 16]
        for col_num, width in enumerate(widths, 1):
            ws.column_dimensions[chr(64 + col_num)].width = width

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="relatorio_financeiro.xlsx"'
        wb.save(response)
        return response


class FinanceExportPdfView(StaffRequiredMixin, View):
    def get(self, request, pk, *args, **kwargs):
        ordem = get_object_or_404(
            OrdemServico.objects.select_related('oficina').prefetch_related('service_items', 'part_items'),
            pk=pk,
            oficina=self.oficina,
        )
        oficina = ordem.oficina
        service_subtotal = sum((item.total_price for item in ordem.service_items.all()), Decimal('0'))
        parts_subtotal = sum((item.total_price for item in ordem.part_items.all()), Decimal('0'))
        discount = getattr(ordem, 'discount', getattr(ordem, 'desconto', Decimal('0'))) or Decimal('0')
        invoice_total = service_subtotal + parts_subtotal - discount
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="cupom_{ordem.order_number}.pdf"'

        from reportlab.lib.pagesizes import A4
        from reportlab.pdfgen import canvas
        from reportlab.lib.units import mm

        buffer = response
        document = canvas.Canvas(buffer, pagesize=A4)
        document.setFont('Helvetica-Bold', 14)
        document.drawString(30 * mm, 280 * mm, oficina.nome or '-')
        document.setFont('Helvetica', 10)
        y = 273 * mm
        office_lines = [
            f'CPF/CNPJ: {oficina.documento}' if oficina.documento else '',
            f'Endereco: {oficina.endereco}' if oficina.endereco else '',
            f'Cidade/UF: {oficina.cidade}/{oficina.estado}' if oficina.cidade or oficina.estado else '',
            f'Telefone/WhatsApp: {oficina.telefone}' if oficina.telefone else '',
            f'E-mail: {oficina.email}' if oficina.email else '',
        ]
        for line in [line for line in office_lines if line]:
            document.drawString(30 * mm, y, line)
            y -= 6 * mm
        document.line(30 * mm, y, 180 * mm, y)
        y -= 8 * mm
        document.setFont('Helvetica-Bold', 12)
        document.drawString(30 * mm, y, f'Cupom de Serviço - {ordem.order_number}')
        y -= 8 * mm
        document.setFont('Helvetica', 10)
        document.drawString(30 * mm, y, f'Data/Hora: {ordem.completed_at.strftime("%d/%m/%Y %H:%M") if ordem.completed_at else "-"}')
        y -= 8 * mm
        document.drawString(30 * mm, y, f'Cliente: {ordem.client_name}')
        y -= 8 * mm
        document.drawString(30 * mm, y, f'Veículo: {ordem.vehicle_brand} {ordem.vehicle_model} - {ordem.plate}')
        y -= 8 * mm
        document.drawString(30 * mm, y, f'Quilometragem: {ordem.mileage or "-"}')
        y -= 12 * mm
        document.setFont('Helvetica-Bold', 11)
        document.drawString(30 * mm, y, 'Serviços executados:')
        y -= 8 * mm
        document.setFont('Helvetica', 10)
        if ordem.service_items.exists():
            for item in ordem.service_items.all():
                document.drawString(32 * mm, y, f'{item.description} x{item.quantity} - R$ {item.unit_price:.2f} un. - R$ {item.total_price:.2f}')
                y -= 7 * mm
                if y < 30 * mm:
                    document.showPage()
                    y = 280 * mm
        else:
            document.drawString(32 * mm, y, '- Nenhum serviço registrado -')
            y -= 7 * mm
        y -= 4 * mm
        document.setFont('Helvetica-Bold', 11)
        document.drawString(30 * mm, y, 'Peças utilizadas:')
        y -= 8 * mm
        document.setFont('Helvetica', 10)
        if ordem.part_items.exists():
            for item in ordem.part_items.all():
                document.drawString(32 * mm, y, f'{item.description} x{item.quantity} - R$ {item.unit_price:.2f} un. - R$ {item.total_price:.2f}')
                y -= 7 * mm
                if y < 30 * mm:
                    document.showPage()
                    y = 280 * mm
        else:
            document.drawString(32 * mm, y, '- Nenhuma peça registrada -')
            y -= 7 * mm
        y -= 8 * mm
        document.setFont('Helvetica-Bold', 11)
        document.drawString(30 * mm, y, f'Total serviços: R$ {service_subtotal:.2f}')
        y -= 7 * mm
        document.drawString(30 * mm, y, f'Total peças: R$ {parts_subtotal:.2f}')
        if discount:
            y -= 7 * mm
            document.drawString(30 * mm, y, f'Desconto: R$ {discount:.2f}')
        y -= 7 * mm
        document.drawString(30 * mm, y, f'Valor total: R$ {invoice_total:.2f}')
        y -= 7 * mm
        document.drawString(30 * mm, y, f'Pagamento: {ordem.get_payment_method_display() if ordem.payment_method else "-"}')
        y -= 7 * mm
        document.drawString(30 * mm, y, f'Mecânico: {ordem.mechanic_name or "-"}')
        y -= 7 * mm
        document.drawString(30 * mm, y, f'Garantia: {ordem.warranty}')
        y -= 12 * mm
        document.setFont('Helvetica', 9)
        document.drawString(30 * mm, y, f'Observações: {ordem.observations or "-"}')
        document.showPage()
        document.save()
        return response


class FinanceInvoiceView(StaffRequiredMixin, TemplateView):
    template_name = 'agenda/finance_invoice.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        os_pk = self.kwargs.get('pk')
        ordem = get_object_or_404(
            OrdemServico.objects.select_related('oficina').prefetch_related('service_items', 'part_items'),
            pk=os_pk,
            oficina=self.oficina,
        )
        service_subtotal = sum((item.total_price for item in ordem.service_items.all()), Decimal('0'))
        parts_subtotal = sum((item.total_price for item in ordem.part_items.all()), Decimal('0'))
        discount = getattr(ordem, 'discount', getattr(ordem, 'desconto', Decimal('0'))) or Decimal('0')
        context.update({
            'ordem': ordem,
            'oficina': ordem.oficina,
            'service_subtotal': service_subtotal,
            'parts_subtotal': parts_subtotal,
            'discount': discount,
            'invoice_total': service_subtotal + parts_subtotal - discount,
        })
        return context


class BookingStatusUpdateView(StaffRequiredMixin, View):
    def post(self, request, *args, **kwargs):
        booking_id = request.POST.get('booking_id')
        new_status = request.POST.get('status')
        booking = get_object_or_404(Booking, pk=booking_id, oficina=self.oficina)
        if new_status in Booking.Status.values:
            booking.status = new_status
            booking.save()
        return redirect(request.META.get('HTTP_REFERER', reverse_lazy('agenda:dashboard')))


class DeleteBookingView(StaffRequiredMixin, View):
    def post(self, request, *args, **kwargs):
        booking_id = request.POST.get('booking_id')
        booking = get_object_or_404(Booking, pk=booking_id, oficina=self.oficina)
        booking.delete()
        return redirect(request.META.get('HTTP_REFERER', reverse_lazy('agenda:dashboard')))
